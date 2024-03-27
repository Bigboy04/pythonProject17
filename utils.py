from io import BytesIO

import openpyxl
import pandas as pd
import math

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


class RaschetStoimosti:
    def __init__(self, type_of_stairs, opening, height_from_floor_to_floor, fence, k, rounding):
        self.type_of_stairs = type_of_stairs
        self.opening = opening
        self.height_from_floor_to_floor = height_from_floor_to_floor
        self.fence = fence
        self.k = k
        self.rounding = rounding
        self.arr = [['Прямая', 'Г - образная (с забежными ступенями)', 'Г - образная (с площадкой)',
                     'П - образная (с забежными ступенями)', 'П - образная (с площадкой)'],
                    ['открытая', 'закрытая'],
                    ['Нет - без ограждений по фаршу', 'Ограждение по 1 стороне марша',
                     'Ограждение по 2-м сторонам марша', 'Ограждение по 1 стороне марша + поручень по стене'],
                    ['Нет', 'Одна пригласительная ступень', 'Две или более пригласительные ступени',
                     'Одна в две стороны', 'Две или более в две стороны']]
        self.arr2 = []

    def create_table_1(self):
        z = self.height_from_floor_to_floor
        x = math.ceil(z / 200)
        y = math.ceil(math.sqrt((x * 300) ** 2 + z ** 2) / 1000)
        tetiva = 2 * y
        stupen = balyasin = stolb = poruchen = ploschadka = opora = podstupen = podbalyasinnik = 0
        round_price = [0, 0, 0, 0]
        coef = 1
        if self.type_of_stairs in [1, 3, 5]:
            stupen = x - 1
            if self.type_of_stairs == 1:
                coef = 2
            else:
                coef = 2.2
        elif self.type_of_stairs == 2:
            stupen = x - 6
            coef = 2.5
        elif self.type_of_stairs == 4:
            stupen = x - 11
            coef = 2.5
        if self.opening == 0:
            podstupen = 0
        elif self.opening == 1:
            podstupen = x
        if self.type_of_stairs == 1:
            ploschadka = 0
            opora = 0
        elif self.type_of_stairs == 2:
            ploschadka = 3.6
            opora = 1
        elif self.type_of_stairs == 3:
            ploschadka = 1
            opora = 4
        elif self.type_of_stairs == 4:
            ploschadka = 7.2
            opora = 1
        elif self.type_of_stairs == 5:
            ploschadka = 2
            opora = 6
        if self.fence == 0:
            balyasin = 0
            stolb = 0
            poruchen = 0
        elif self.fence == 1:
            balyasin = x - 1
            stolb = 2
            poruchen = y
        elif self.fence == 2:
            balyasin = 2 * x - 2
            stolb = 4
            poruchen = 2 * y
        elif self.fence == 3:
            balyasin = x - 1
            stolb = 2
            poruchen = 2 * y
        if self.k != 0:
            poruchen += self.k
            podbalyasinnik = self.k
            balyasin = math.ceil(balyasin + 5 * self.k)
            stolb += 2
        if self.rounding == 1:
            round_price = [7000, 8000, 11000, 20000]
        elif self.rounding == 2:
            round_price = [14000, 16000, 22000, 40000]
        elif self.rounding == 3:
            round_price = [10500, 12000, 16500, 30000]
        elif self.rounding == 4:
            round_price = [21000, 24000, 33000, 60000]
        kolichestvo = [stupen, balyasin, stolb, poruchen, tetiva, podstupen, ploschadka, opora, podbalyasinnik]
        df = pd.read_excel('Книга.xlsx', index_col=0)
        df.insert(0, 'Кол-во', kolichestvo)

        for col_index in range(1, len(df.columns)):
            s = 0
            for row_index in range(len(df)):
                cell_value = df.iloc[row_index, col_index]
                s += cell_value * df.iloc[row_index, 0]
            self.arr2.append(s*coef + round_price[col_index-1])

        output_file = BytesIO()
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer)
        output_file.seek(0)

        wb = load_workbook(output_file)
        ws = wb.active
        row = ws.max_row

        border_style1 = Border(left=Side(style='thin'), right=Side(style='thin'))
        for r in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col)
                if col != 1 and r != 1:
                    ws[f'{column_letter}{r}'].border = border_style1
                else:
                    ws[f'{column_letter}{r}'].border = Border(left=Side(style='medium'), right=Side(style='medium'),
                                                              top=Side(style='medium'), bottom=Side(style='medium'))

        for col in range(3, ws.max_column + 1):
            column_letter = get_column_letter(col)
            for i in range(2, row + 1):
                value1 = ws[f'{column_letter}{i}'].value
                ws[f'{column_letter}{i}'] = f'={value1}*B{i}'

        for col in range(3, ws.max_column + 1):
            column_letter = get_column_letter(col)
            formula = f"=SUM({column_letter}2:{column_letter}{row})"
            ws[f"{column_letter}{row + 1}"] = formula

        ws[f'B{row + 1}'] = "Материал"
        ws[f'B{row + 2}'] = "ИТОГО"

        border_style2 = Border(left=Side(style='medium'),
                               right=Side(style='medium'),
                               top=Side(style='medium'),
                               bottom=Side(style='medium'))

        for col in range(3, ws.max_column + 1):
            column_letter = get_column_letter(col)
            ws[f"{column_letter}{row + 2}"] = f"={column_letter}{row + 1}*{coef} + {round_price[col - 3]}"

        for col in range(2, ws.max_column + 1):
            for r in range(row + 1, row + 3):
                column_letter = get_column_letter(col)
                ws[f"{column_letter}{r}"].font = Font(bold=True)
                ws[f"{column_letter}{r}"].border = border_style2

        ws[f'A{row + 5}'] = self.arr[0][self.type_of_stairs - 1]
        ws[f'A{row + 6}'] = self.arr[1][self.opening]
        ws[f'A{row + 7}'] = self.height_from_floor_to_floor
        ws[f'A{row + 8}'] = self.arr[2][self.fence]
        ws[f'A{row + 9}'] = self.k
        ws[f'A{row + 10}'] = self.arr[3][self.rounding]

        output_file2 = BytesIO()
        wb.save(output_file2)
        output_file2.seek(0)
        return output_file2
